import openpyxl
import webbrowser
import os


def read(file_name):
    my_xlsx = openpyxl.load_workbook(file_name)
    my_sheet = my_xlsx.active
    xlsx_to_list =[]

    for row in my_sheet.iter_rows():
        xlsx_to_list_row = []
        for cell in row:
            xlsx_to_list_row.append(cell.value)
        xlsx_to_list.append(xlsx_to_list_row)
    my_xlsx.close()
    return xlsx_to_list


def write(file_name, data_list):
    workbook= openpyxl.Workbook()
    my_xlsx = workbook.active
    my_xlsx.sheet_view.rightToLeft = True

    for row in data_list:
        my_xlsx.append(row)
    workbook.save(filename=file_name)
    workbook.close()
    print('completed')


def append(file_name, sheet_name, data_list):

    try:
        my_xlsx = openpyxl.load_workbook(file_name)
    except:
        print("Cant find the file:  " + file_name)

    else:
        worksheet = my_xlsx.create_sheet(title=sheet_name)

        for row in data_list:
            worksheet.append(row)
        worksheet.sheet_view.rightToLeft = True
        my_xlsx.save(filename=file_name)
        print('completed')


def view(file_name):

    cwd = os.getcwd()
    file_path = os.path.join(cwd, file_name)
    try:
        webbrowser.open('file://' + file_path)
    except FileExistsError:
        print("Cant open the file:  " + file_name)
